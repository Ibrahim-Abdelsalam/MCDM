"""Excel report generation for MCDM analysis results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Fill, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from ..config import BASE_DIR


class ExcelExporter:
    """Export analysis results to formatted Excel workbooks."""

    def __init__(self, output_dir: Path | None = None) -> None:
        """Initialize the exporter with the output directory."""
        self.output_dir = output_dir or (BASE_DIR / "data" / "results")

    def export(
        self,
        entity_type: str,
        analysis_results: dict[str, pd.DataFrame],
    ) -> Path:
        """
        Export analysis results to an Excel workbook.
        
        Args:
            entity_type: "supplier" or "buyer"
            analysis_results: dict mapping "{METHOD}_{SCHEME}" to result DataFrame
        
        Returns:
            Path to the generated Excel file
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{entity_type}_results_{timestamp}.xlsx"
        file_path = self.output_dir / filename
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create one sheet per method-scheme combination
        for key, result_df in analysis_results.items():
            sheet_name = key[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            
            # Prepare data: Rank, ID, Score (sorted by rank)
            sorted_df = result_df.sort_values("Rank").reset_index(drop=True)
            
            # Write header
            ws.append(["Rank", "Company_ID", "Score"])
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Write data rows
            top_10_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            for row_idx, (_, row) in enumerate(sorted_df.iterrows(), start=2):
                ws.append([
                    int(row["Rank"]),
                    str(row["ID"]),
                    float(row["Score"]),
                ])
                
                if row_idx <= 11:  # Top 10 + header
                    for cell in ws[row_idx]:
                        cell.fill = top_10_fill
            
            # Set column widths
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 12
        
        # Create summary sheet with top 5 from each method-scheme
        summary_data = []
        for key, result_df in analysis_results.items():
            method, scheme = key.rsplit("_", 1)
            top_5 = result_df.nsmallest(5, "Rank")
            for _, row in top_5.iterrows():
                summary_data.append({
                    "Method": method,
                    "Weight_Scheme": scheme,
                    "Rank": int(row["Rank"]),
                    "Company_ID": str(row["ID"]),
                    "Score": float(row["Score"]),
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            ws_summary = wb.create_sheet(title="Summary_Top5", index=0)
            
            # Write header
            ws_summary.append(summary_df.columns.tolist())
            
            summary_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            summary_header_font = Font(bold=True, color="FFFFFF")
            summary_header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws_summary[1]:
                cell.fill = summary_header_fill
                cell.font = summary_header_font
                cell.alignment = summary_header_alignment
            
            # Write data
            for _, row in summary_df.iterrows():
                ws_summary.append([
                    row["Method"],
                    row["Weight_Scheme"],
                    row["Rank"],
                    row["Company_ID"],
                    row["Score"],
                ])
            
            # Set column widths
            for col in ws_summary.columns:
                ws_summary.column_dimensions[col[0].column_letter].width = 15
        
        wb.save(file_path)
        return file_path
